from openpyxl import load_workbook
from local import localPath
import utility

month = utility.getLastMonth()
year = utility.getLastMonthYear()
clientes = []

controlExcelPath = f"{localPath}\\CONTROL\\Control.xlsx"

def setClientes():
    workbook = load_workbook(controlExcelPath)
    firstSheet = workbook.active
    lastRow = firstSheet.max_row

    for i in range(1, lastRow):

        # Obtener los datos de esa fila
        categoria = firstSheet["A" + str(i + 1)].value
        rfc = firstSheet["B" + str(i + 1)].value
        usuario = firstSheet["C" + str(i + 1)].value
        contrasenia = firstSheet["D" + str(i + 1)].value
        ruta = f"{localPath}\\DESCARGAS\\{rfc}\\{str(year)}\\{str(month)}"

        clientes.append({
            "categoria" : categoria,
            "rfc" : rfc,
            "usuario" : usuario,
            "contraseña" : contrasenia,
            "ruta" : ruta
        })